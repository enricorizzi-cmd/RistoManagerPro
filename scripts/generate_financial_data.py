# -*- coding: utf-8 -*-
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(r"c:/RistoManagerPro")
DATA_DIR = ROOT / "data"
DATA_DIR.mkdir(exist_ok=True)


def build_financial_plan() -> list[dict[str, Any]]:
    workbook = load_workbook(ROOT / "Piano Finanziario - FP/PIANO FINANZIARIO.xlsx", data_only=True)
    sheet = workbook.active

    months: list[dict[str, Any]] = []
    col = 3
    max_col = sheet.max_column or 0
    while col <= max_col:
        month_name = sheet.cell(row=1, column=col).value
        if month_name:
            months.append({
                "label": str(month_name).strip(),
                "preventivo_col": col,
                "consuntivo_col": col + 1,
            })
        col += 2

    rows: list[dict[str, Any]] = []
    current_macro: str | None = None
    for raw in sheet.iter_rows(min_row=3, values_only=True):
        macro = raw[0] or current_macro
        if macro:
            current_macro = str(macro).strip()
        detail = str(raw[1]).strip() if raw[1] else ""

        month_values = []
        for month in months:
            prev_val = raw[month["preventivo_col"] - 1]
            cons_val = raw[month["consuntivo_col"] - 1]
            prev = float(prev_val) if isinstance(prev_val, (int, float)) else None
            cons = float(cons_val) if isinstance(cons_val, (int, float)) else None
            month_values.append({
                "month": month["label"],
                "preventivo": prev,
                "consuntivo": cons,
            })

        if not (current_macro or detail):
            continue
        if detail == "" and all(v["preventivo"] is None and v["consuntivo"] is None for v in month_values):
            continue

        rows.append({
            "macroCategory": current_macro or "",
            "detail": detail,
            "months": month_values,
        })

    return rows


def build_causali() -> list[dict[str, Any]]:
    workbook = load_workbook(ROOT / "Piano Finanziario - FP/Causali FP.xlsx", data_only=True)
    sheet = workbook.active

    grouped: dict[str, dict[str, list[str]]] = {}
    for macro, category, item in sheet.iter_rows(values_only=True):
        if not macro:
            continue
        macro_key = str(macro).strip()
        category_key = str(category).strip() if category else "Generale"
        item_value = str(item).strip() if item else None
        category_map = grouped.setdefault(macro_key, {})
        items_list = category_map.setdefault(category_key, [])
        if item_value and item_value not in items_list:
            items_list.append(item_value)

    result: list[dict[str, Any]] = []
    for macro_key in sorted(grouped.keys()):
        categories = grouped[macro_key]
        result.append({
            "macroCategory": macro_key,
            "categories": [
                {
                    "name": category_name,
                    "items": sorted(items),
                }
                for category_name, items in sorted(categories.items())
            ],
        })
    return result


def _parse_number(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace("\u20ac", "").replace(" ", "").replace(".", "").replace(",", ".").replace("-", "")
        if cleaned == "":
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def build_stats() -> list[dict[str, Any]]:
    workbook = load_workbook(ROOT / "Piano Finanziario - FP/STATISTICHE FP.xlsx", data_only=True)
    sheet = workbook.active

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    sanitize = lambda h: str(h).replace("\n", " ").strip()
    header_map = {sanitize(header_row[idx]): idx for idx in range(len(header_row)) if header_row[idx]}

    target_headers = {
        "Valore Fatturato IMPONIBILE": "fatturatoImponibile",
        "Valore Fatturato + Corrispettivi": "fatturatoTotale",
        "UTILE DI CASSA": "utileCassa",
        "Valore Incassato": "incassato",
        "Saldo conto corrente": "saldoConto",
        "Saldo Secondo Conto": "saldoSecondoConto",
        "Somma Saldo Conti": "saldoTotale",
        "Valore Crediti pendenti": "creditiPendenti",
        "Valore Crediti Scaduti": "creditiScaduti",
        "Debiti Fornitore": "debitiFornitore",
        "Debiti Bancari (leasing compreso)": "debitiBancari",
    }

    rows: list[dict[str, Any]] = []
    for raw in sheet.iter_rows(min_row=2, values_only=True):
        month = raw[0]
        if not month:
            continue
        entry: dict[str, Any] = {"month": str(month).strip()}
        has_value = False
        for header, key in target_headers.items():
            idx = header_map.get(header)
            value = _parse_number(raw[idx]) if idx is not None else None
            entry[key] = value
            if value is not None:
                has_value = True
        if has_value:
            rows.append(entry)

    return rows


def main() -> None:
    plan_rows = build_financial_plan()
    causali_rows = build_causali()
    stats_rows = build_stats()

    ts_header = """// This file is auto-generated from the Excel sources in Piano Finanziario - FP
// Run python scripts/generate_financial_data.py to refresh the dataset.

export interface FinancialPlanMonthValue {
  month: string;
  preventivo: number | null;
  consuntivo: number | null;
}

export interface FinancialPlanRow {
  macroCategory: string;
  detail: string;
  months: FinancialPlanMonthValue[];
}

export interface FinancialCausaleCategory {
  name: string;
  items: string[];
}

export interface FinancialCausaleGroup {
  macroCategory: string;
  categories: FinancialCausaleCategory[];
}

export interface FinancialStatsRow {
  month: string;
  fatturatoImponibile: number | null;
  fatturatoTotale: number | null;
  utileCassa: number | null;
  incassato: number | null;
  saldoConto: number | null;
  saldoSecondoConto: number | null;
  saldoTotale: number | null;
  creditiPendenti: number | null;
  creditiScaduti: number | null;
  debitiFornitore: number | null;
  debitiBancari: number | null;
}

"""

    def to_ts(name: str, data: Any) -> str:
        json_block = json.dumps(data, ensure_ascii=False, indent=2)
        return f"export const {name} = {json_block} as const;\n"

    ts_body = "\n".join([
        ts_header,
        to_ts("financialPlanRows", plan_rows),
        "",
        to_ts("financialCausali", causali_rows),
        "",
        to_ts("financialStats", stats_rows),
    ])

    target_file = DATA_DIR / "financialPlanData.ts"
    target_file.write_text(ts_body, encoding="utf-8")
    print(f"Wrote {target_file.relative_to(ROOT)}")


if __name__ == "__main__":
    main()

